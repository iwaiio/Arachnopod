#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import math
import re
import sys

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
SOFTWARE_DIR = SCRIPT_DIR.parent
EXCEL_PATH = SOFTWARE_DIR / "data_base" / "BD.xlsx"
OUTPUT_DIR = SOFTWARE_DIR / "sim" / "sim" / "systems" / "base"
SIM_BASE_INCLUDE_DIR = SOFTWARE_DIR / "sim" / "sim" / "simulation" / "base" / "include"

LINE_WIDTH = 140
BLOCK_BITS = 16
TAR_NONE = 0
LOCAL_PARAM_ID_BASE = 0x8000


def fail(message: str) -> None:
    print(f"error: {message}", file=sys.stderr)
    raise SystemExit(1)


def warn(message: str) -> None:
    print(f"warning: {message}")


def error_log(message: str) -> None:
    print(f"error: {message}", file=sys.stderr)


def info(message: str) -> None:
    print(f"info: {message}")


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("*", "").replace(" ", "_")
    return text


def sanitize_identifier(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = text.replace("Ё", "Е")
    text = re.sub(r"[^A-Z0-9_]", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = "UNNAMED"
    if text[0].isdigit():
        text = f"_{text}"
    return text


def c_comment(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.replace("/*", "/ *").replace("*/", "* /")


def to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return int(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_cpp_number(value: Any) -> str:
    if value in (None, ""):
        return "0"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).replace(",", ".")
    text = str(value).strip().replace(",", ".")
    return text if text else "0"


def pad140(line: str) -> str:
    if len(line) >= LINE_WIDTH:
        return line
    return line + (" " * (LINE_WIDTH - len(line)))


def banner(source: str) -> str:
    rows = [
        "/* -------------------------------------------------------------------------------------------------------------------------------------- */",
        f"/* AUTO-GENERATED FILE. Source: {source:<105} */",
        "/* Do not edit manually.                                                                                                                  */",
        "/* -------------------------------------------------------------------------------------------------------------------------------------- */",
        "",
    ]
    return "\n".join(rows)


def build_header_map(headers: list[Any]) -> dict[str, int]:
    return {normalize_header(h): i for i, h in enumerate(headers)}


def get_by_header(row: list[Any], header_map: dict[str, int], *names: str, default: Any = None) -> Any:
    for name in names:
        idx = header_map.get(normalize_header(name))
        if idx is not None and idx < len(row):
            return row[idx]
    return default

def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def type_size_bits(type_name: str) -> int:
    normalized = sanitize_identifier(type_name)
    if normalized == "D":
        return 1
    if normalized == "A":
        return 8
    if normalized == "AP":
        return 16
    fail(f"Неизвестный тип параметра: {type_name}")
    return 0


def validate_block_layout(rows: list[BaseRow], label: str) -> None:
    for row in rows:
        if row.is_placeholder:
            continue

        width = type_size_bits(row.type_name)
        offset = row.msg_block_offset

        if row.msg_block_n < 0:
            fail(f"{label}: {row.name} has negative MSG_Block_N={row.msg_block_n}")

        if width == 16:
            if offset != 0:
                fail(f"{label}: {row.name} type AP must start at bit 0, got MSG_Block_Offset={offset}")
            continue

        if width == 8:
            if offset not in (0, 8):
                fail(f"{label}: {row.name} type A must start at bit 0 or 8, got MSG_Block_Offset={offset}")
            continue

        if width == 1:
            if offset < 0 or offset >= BLOCK_BITS:
                fail(f"{label}: {row.name} type D must be in bit range [0, 15], got MSG_Block_Offset={offset}")
            continue

        fail(f"{label}: unsupported width={width} for {row.name}")


def sign_macro(sign: str) -> str:
    return "TYPE_SIGN" if sanitize_identifier(sign) == "S" else "TYPE_UNSIGN"


def make_aligned_comment(left: str, right: str = "", width: int = LINE_WIDTH) -> str:
    left_text = c_comment(left)
    right_text = c_comment(right)
    inner_width = max(0, width - len("/** ") - len(" */"))
    if right_text:
        spacer = max(1, inner_width - len(left_text) - len(right_text))
        body = f"{left_text}{' ' * spacer}{right_text}"
    else:
        body = left_text
    if len(body) < inner_width:
        body = body + (" " * (inner_width - len(body)))
    else:
        body = body[:inner_width]
    return f"/** {body} */"


def field_line(c_type: str, name: str, right_comment: str) -> str:
    prefix = f"	{c_type:<10}	{name:<20}	"
    comment = make_aligned_comment(name.strip(';'), right_comment, LINE_WIDTH - len(prefix))
    return (prefix + comment)[:LINE_WIDTH]


@dataclass
class BaseRow:
    row_num: int
    id: int
    name: str
    description: str
    system_name: str
    type_name: str
    type_real: str
    sign: str
    tar_a: Any
    tar_b: Any
    tar_c: Any
    alg_name: str
    msg_offset: int
    msg_block_offset: int
    msg_block_n: int
    init_value: float = 0.0
    msg_cs_offset: int = 0
    msg_cs_block_n: int = 0
    is_placeholder: bool = False


@dataclass
class ParamRow(BaseRow):
    pass


@dataclass
class CommandRow(BaseRow):
    pass


@dataclass
class AlgorithmRow:
    row_num: int
    name: str
    encode_expr: str
    decode_expr: str
    condition_expr: str


@dataclass
class ModelRow:
    row_num: int
    id: int | None
    name: str
    system_name: str
    type_name: str
    type_real: str
    sign: str
    init_value: float = 0.0


ModelLike = BaseRow | ModelRow


@dataclass
class SystemStats:
    blocks: int
    bits: int


def load_data_book(path: Path):
    if not path.exists():
        fail(f"Excel-файл не найден: {path}")
    return load_workbook(path, data_only=True)


def parse_sheet(wb, sheet_name: str, kind: str):
    if sheet_name not in wb.sheetnames:
        fail(f"В книге отсутствует лист '{sheet_name}'")

    sheet = wb[sheet_name]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    hm = build_header_map(headers)

    raw_rows: list[dict[str, Any]] = []
    used_ids: set[int] = set()

    for r in range(2, sheet.max_row + 1):
        data = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]

        raw_id = get_by_header(data, hm, "ID")
        row_id = to_int(raw_id)

        raw_name = get_by_header(data, hm, "Name", default="")
        raw_description = get_by_header(data, hm, "Description", default="")
        raw_system = get_by_header(data, hm, "System", default="NONE")
        raw_type = get_by_header(data, hm, "Type", default="D")
        raw_type_real = get_by_header(data, hm, "TypeR", "Type_R")
        raw_sign = get_by_header(data, hm, "Syg", "Sign", default="U")
        raw_tar_a = get_by_header(data, hm, "tar_A")
        raw_tar_b = get_by_header(data, hm, "tar_B")
        raw_tar_c = get_by_header(data, hm, "tar_C")
        raw_alg = get_by_header(data, hm, "Alg", "ALG", default="NONE")
        raw_msg_offset = get_by_header(data, hm, "MSG_Offset")
        raw_msg_block_offset = get_by_header(data, hm, "MSG_Block_Offset")
        raw_msg_block_n = get_by_header(data, hm, "MSG_Block_N")
        raw_init_value = get_by_header(data, hm, "val", "Val", default=None)

        all_empty = all(
            x in (None, "")
            for x in [
                raw_id,
                raw_name,
                raw_description,
                raw_system,
                raw_type,
                raw_type_real,
                raw_sign,
                raw_tar_a,
                raw_tar_b,
                raw_tar_c,
                raw_alg,
                raw_msg_offset,
                raw_msg_block_offset,
                raw_msg_block_n,
                raw_init_value,
            ]
        )
        if all_empty:
            continue

        if row_id is not None:
            if row_id in used_ids:
                fail(f"{kind}: обнаружен повторяющийся ID={row_id} (строка {r})")
            used_ids.add(row_id)

        name = str(raw_name or "").strip()
        if not name:
            warn(f"{kind}: строка {r}, отсутствует Name -> будет NONE_<ID>")

        description = str(raw_description or "")
        if raw_description in (None, ""):
            warn(f"{kind}: строка {r}, отсутствует Description -> подставлено пустое значение")

        system_name = str(raw_system or "NONE").strip() or "NONE"
        if raw_system in (None, ""):
            warn(f"{kind}: строка {r}, отсутствует System -> подставлено NONE")

        type_name = str(raw_type or "D").strip().upper() or "D"
        if raw_type in (None, ""):
            warn(f"{kind}: строка {r}, отсутствует Type -> подставлено D")

        type_real = str(raw_type_real or raw_sign or "U").strip().upper() or "U"
        if raw_type_real in (None, ""):
            warn(f"{kind}: row {r}, missing TypeR -> fallback to Sign/Syg")

        sign = str(raw_sign or "U").strip().upper() or "U"
        if raw_sign in (None, ""):
            warn(f"{kind}: строка {r}, отсутствует Sign/Syg -> подставлено U")

        tar_a = raw_tar_a
        tar_b = raw_tar_b
        tar_c = raw_tar_c
        if tar_a in (None, ""):
            warn(f"{kind}: строка {r}, отсутствует tar_A -> подставлен TAR_NONE=0")
            tar_a = TAR_NONE
        if tar_b in (None, ""):
            warn(f"{kind}: строка {r}, отсутствует tar_B -> подставлен TAR_NONE=0")
            tar_b = TAR_NONE
        if tar_c in (None, ""):
            warn(f"{kind}: строка {r}, отсутствует tar_C -> подставлен TAR_NONE=0")
            tar_c = TAR_NONE

        alg_name = str(raw_alg or "NONE").strip() or "NONE"
        if raw_alg in (None, ""):
            warn(f"{kind}: строка {r}, отсутствует Alg -> подставлено NONE")

        msg_offset = to_int(raw_msg_offset)
        msg_block_offset = to_int(raw_msg_block_offset)
        msg_block_n = to_int(raw_msg_block_n)
        if msg_offset is None:
            warn(f"{kind}: строка {r}, отсутствует MSG_Offset -> подставлено 0")
            msg_offset = 0
        if msg_block_offset is None:
            warn(f"{kind}: строка {r}, отсутствует MSG_Block_Offset -> подставлено 0")
            msg_block_offset = 0
        if msg_block_n is None:
            warn(f"{kind}: строка {r}, отсутствует MSG_Block_N -> подставлено 0")
            msg_block_n = 0

        type_real = canonical_physical_type(type_real, type_name, sign, f"{kind}: row {r} name={name or row_id}")

        raw_rows.append(
            {
                "row_num": r,
                "id": row_id,
                "name": name,
                "description": description,
                "system_name": system_name,
                "type_name": type_name,
                "type_real": type_real,
                "sign": sign,
                "tar_a": tar_a,
                "tar_b": tar_b,
                "tar_c": tar_c,
                "alg_name": alg_name,
                "msg_offset": msg_offset,
                "msg_block_offset": msg_block_offset,
                "msg_block_n": msg_block_n,
                "init_value": 0.0 if to_float(raw_init_value) is None else to_float(raw_init_value),
            }
        )

    if not raw_rows:
        fail(f"Лист '{sheet_name}' пуст")

    next_id = min(used_ids) if used_ids else 0

    def allocate_id() -> int:
        nonlocal next_id
        while next_id in used_ids:
            next_id += 1
        value = next_id
        used_ids.add(value)
        next_id += 1
        return value

    rows: list[BaseRow] = []
    for item in raw_rows:
        row_id = item["id"]
        if row_id is None:
            row_id = allocate_id()
            error_log(
                f"{kind}: строка {item['row_num']}, отсутствует ID -> присвоен автоматически ID={row_id}"
            )

        name = item["name"]
        if not name:
            name = f"NONE_{row_id}"

        rows.append(
            BaseRow(
                row_num=item["row_num"],
                id=row_id,
                name=name,
                description=item["description"],
                system_name=item["system_name"],
                type_name=item["type_name"],
                type_real=item["type_real"],
                sign=item["sign"],
                tar_a=item["tar_a"],
                tar_b=item["tar_b"],
                tar_c=item["tar_c"],
                alg_name=item["alg_name"],
                msg_offset=item["msg_offset"],
                msg_block_offset=item["msg_block_offset"],
                msg_block_n=item["msg_block_n"],
                init_value=item["init_value"],
            )
        )

    return rows


def parse_model_sheet(wb, sheet_name: str, kind: str) -> list[ModelRow]:
    if sheet_name not in wb.sheetnames:
        warn(f"В книге отсутствует лист '{sheet_name}'")
        return []

    sheet = wb[sheet_name]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    hm = build_header_map(headers)

    raw_rows: list[dict[str, Any]] = []
    used_ids: set[int] = set()

    for r in range(2, sheet.max_row + 1):
        data = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]

        raw_id = get_by_header(data, hm, "ID")
        row_id = to_int(raw_id)

        raw_name = get_by_header(data, hm, "Name", default="")
        raw_system = get_by_header(data, hm, "System", default="NONE")
        raw_type = get_by_header(data, hm, "Type", default="D")
        raw_type_real = get_by_header(data, hm, "TypeR", "Type_R")
        raw_sign = get_by_header(data, hm, "Syg", "Sign", default="U")
        raw_init_value = get_by_header(data, hm, "val", "Val", default=None)

        all_empty = all(
            x in (None, "")
            for x in [
                raw_id,
                raw_name,
                raw_system,
                raw_type,
                raw_type_real,
                raw_sign,
                raw_init_value,
            ]
        )
        if all_empty:
            continue

        if row_id is not None:
            if row_id in used_ids:
                fail(f"{kind}: обнаружен повторяющийся ID={row_id} (строка {r})")
            used_ids.add(row_id)

        raw_rows.append(
            {
                "row_num": r,
                "id": row_id,
                "name": str(raw_name or "").strip(),
                "system_name": str(raw_system or "NONE").strip() or "NONE",
                "type_name": str(raw_type or "D").strip().upper() or "D",
                "type_real": str(raw_type_real or "").strip().upper(),
                "sign": str(raw_sign or "U").strip().upper() or "U",
                "init_value": 0.0 if to_float(raw_init_value) is None else to_float(raw_init_value),
            }
        )

    if not raw_rows:
        warn(f"Лист '{sheet_name}' пуст")
        return []

    rows: list[ModelRow] = []
    for item in raw_rows:
        name = item["name"]
        if not name:
            fallback_id = item["id"] if item["id"] is not None else 0
            name = f"NONE_{fallback_id}"
        elif kind == "Local_Params" and not sanitize_identifier(name).startswith("L"):
            fail(f"{kind}: row {item['row_num']} name={name} must start with 'L'")

        type_name = item["type_name"]
        if not type_name:
            warn(f"{kind}: строка {item['row_num']}, отсутствует Type -> подставлено D")
            type_name = "D"

        type_real = item["type_real"]
        if not type_real:
            type_real = item["sign"] or "U"

        sign = item["sign"]
        if not sign:
            warn(f"{kind}: строка {item['row_num']}, отсутствует Sign/Syg -> подставлено U")
            sign = "U"

        type_real = canonical_physical_type(type_real, type_name, sign, f"{kind}: row {item['row_num']} name={name or row_id}")

        rows.append(
            ModelRow(
                row_num=item["row_num"],
                id=item["id"],
                name=name,
                system_name=item["system_name"],
                type_name=type_name,
                type_real=type_real,
                sign=sign,
                init_value=item["init_value"],
            )
        )

    return rows


def parse_algorithm_sheet(wb, sheet_name: str) -> list[AlgorithmRow]:
    if sheet_name not in wb.sheetnames:
        fail(f"В книге отсутствует лист '{sheet_name}'")

    sheet = wb[sheet_name]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    hm = build_header_map(headers)
    rows: list[AlgorithmRow] = []
    used_names: set[str] = set()

    for r in range(2, sheet.max_row + 1):
        data = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
        raw_name = get_by_header(data, hm, "Alg", "Name", default="")
        raw_encode = get_by_header(data, hm, "encode", "Encode", default="")
        raw_decode = get_by_header(data, hm, "decode", "Decode", default="")
        raw_condition = get_by_header(data, hm, "if", "If", "Condition", default="")

        if all(x in (None, "") for x in [raw_name, raw_encode, raw_decode, raw_condition]):
            continue

        name = str(raw_name or "").strip().upper()
        if not name:
            fail(f"Algorithms: строка {r}, отсутствует Alg")

        key = sanitize_identifier(name)
        if key in used_names:
            fail(f"Algorithms: обнаружен повторяющийся алгоритм {name} (строка {r})")
        used_names.add(key)

        rows.append(
            AlgorithmRow(
                row_num=r,
                name=name,
                encode_expr=str(raw_encode or "").strip(),
                decode_expr=str(raw_decode or "").strip(),
                condition_expr=str(raw_condition or "").strip(),
            )
        )

    if not rows:
        fail(f"Лист '{sheet_name}' пуст")

    return rows


def canonical_physical_type(type_real: str, type_name: str, sign: str, row_label: str) -> str:
    key = sanitize_identifier(type_real)
    width = type_size_bits(type_name)

    if key in ("BOOL", "BOOLEAN"):
        if width != 1:
            fail(f"{row_label}: bool physical type must use wire type D (1 bit), got {type_name}")
        return "BOOL"

    if width == 1:
        if not key:
            return "BOOL"
        fail(f"{row_label}: 1-bit wire type D only supports physical type bool, got {type_real}")

    if key in ("INT8", "I8"):
        if width != 8:
            fail(f"{row_label}: int8 physical type requires wire type A (8 bit), got {type_name}")
        return "INT8"

    if key in ("UINT8", "U8"):
        if width != 8:
            fail(f"{row_label}: uint8 physical type requires wire type A (8 bit), got {type_name}")
        return "UINT8"

    if key in ("INT16", "I16"):
        if width != 16:
            fail(f"{row_label}: int16 physical type requires wire type AP (16 bit), got {type_name}")
        return "INT16"

    if key in ("UINT16", "U16"):
        if width != 16:
            fail(f"{row_label}: uint16 physical type requires wire type AP (16 bit), got {type_name}")
        return "UINT16"

    if key in ("FLOAT", "FLOAT32", "F32", "SFLOAT", "FS"):
        return "FLOAT32"

    if key in ("UFLOAT", "UFLOAT32", "UF32", "FU"):
        return "UFLOAT32"

    if key in ("INT", "SIGNED", "S", "I", "SI"):
        return "INT16" if width == 16 else "INT8"

    if key in ("UINT", "UNSIGNED", "U", "UI", ""):
        if sanitize_identifier(sign) == "S":
            return "INT16" if width == 16 else "INT8"
        return "UINT16" if width == 16 else "UINT8"

    fail(f"{row_label}: unsupported physical type '{type_real}'")
    return "UINT8"


def fill_id_gaps(rows: list[BaseRow], kind: str):
    ids = [r.id for r in rows]
    if len(ids) != len(set(ids)):
        fail(f"{kind}: обнаружены повторяющиеся ID")

    max_id = max(ids)
    by_id = {r.id: r for r in rows}
    result: list[BaseRow] = []

    for idx in range(0, max_id + 1):
        if idx in by_id:
            result.append(by_id[idx])
            continue

        warn(f"{kind}: отсутствует ID={idx}, вставлен резервный NONE")
        result.append(
            BaseRow(
                row_num=0,
                id=idx,
                name=f"NONE_{idx}",
                description="RESERVED",
                system_name="NONE",
                type_name="D",
                type_real="U",
                sign="U",
                tar_a=TAR_NONE,
                tar_b=TAR_NONE,
                tar_c=TAR_NONE,
                alg_name="NONE",
                msg_offset=0,
                msg_block_offset=0,
                msg_block_n=0,
                init_value=0.0,
                is_placeholder=True,
            )
        )

    return result


def cast_params(rows: list[BaseRow]) -> list[ParamRow]:
    return [ParamRow(**r.__dict__) for r in rows]


def cast_commands(rows: list[BaseRow]) -> list[CommandRow]:
    return [CommandRow(**r.__dict__) for r in rows]


def ordered_unique(values: list[str], head: str = "NONE") -> list[str]:
    seen: set[str] = set()
    out: list[str] = []

    def add(v: str) -> None:
        k = sanitize_identifier(v)
        if k in seen:
            return
        seen.add(k)
        out.append(v)

    add(head)
    for value in values:
        if sanitize_identifier(value) != sanitize_identifier(head):
            add(value)
    return out


def collect_systems(params: list[ParamRow], commands: list[CommandRow]) -> list[str]:
    return ordered_unique([*(x.system_name for x in params), *(x.system_name for x in commands)], "NONE")


def collect_algs(algorithms: list[AlgorithmRow]) -> list[str]:
    return ordered_unique([x.name for x in algorithms], "NONE")

def collect_types(params: list[ParamRow], commands: list[CommandRow]) -> list[str]:
    base = ["NONE", "D", "A", "AP", "SIGN", "UNSIGN"]
    used = {sanitize_identifier(x) for x in base}
    values = {sanitize_identifier(x.type_name): x.type_name for x in [*params, *commands]}
    out = list(base)
    for _, value in sorted(values.items(), key=lambda kv: kv[0]):
        key = sanitize_identifier(value)
        if key not in used:
            out.append(value)
            used.add(key)
    return out


def validate_algorithm_references(rows: list[BaseRow], algorithms: list[AlgorithmRow], label: str) -> None:
    known = {sanitize_identifier(x.name) for x in algorithms}
    known.add("NONE")

    for row in rows:
        alg_key = sanitize_identifier(row.alg_name)
        if alg_key not in known:
            fail(f"{label}: {row.name} references unknown algorithm '{row.alg_name}'")


def validate_algorithm_conditions(rows: list[BaseRow], label: str) -> None:
    for row in rows:
        alg_key = sanitize_identifier(row.alg_name)
        row_label = f"{label}: {row.name}"

        if alg_key == "ABC" and to_float(row.tar_b) == 0.0:
            fail(f"{row_label} uses ABC but tar_B == 0")

        if alg_key == "AAC" and to_float(row.tar_a) == 0.0:
            fail(f"{row_label} uses AAC but tar_A == 0")


def report_quantization_errors(rows: list[BaseRow], label: str) -> None:
    for row in rows:
        alg_key = sanitize_identifier(row.alg_name)
        error: float | None = None

        if alg_key == "ABC":
            b = to_float(row.tar_b)
            if b is not None:
                error = abs(b) / 2.0
        elif alg_key == "AAC":
            warn(f"{label}: {row.name} uses AAC, quantization error is not constant and is not reported")

        if error is None or error == 0.0:
            continue

        info(
            f"{label}: {row.name} alg={sanitize_identifier(row.alg_name)} "
            f"max quantization error = {error:g}"
        )


def system_local_bits(rows: list[BaseRow], system_key: str) -> tuple[int, list[tuple[int, int, BaseRow]]]:
    local_rows = [r for r in rows if sanitize_identifier(r.system_name) == system_key and not r.is_placeholder]
    if not local_rows:
        return 0, []

    min_block = min(r.msg_block_n for r in local_rows)
    intervals: list[tuple[int, int, BaseRow]] = []
    max_end = 0

    for row in local_rows:
        size = type_size_bits(row.type_name)
        start = (row.msg_block_n - min_block) * BLOCK_BITS + row.msg_block_offset
        end = start + size
        max_end = max(max_end, end)
        intervals.append((start, end, row))

    blocks = max(1, math.ceil(max_end / BLOCK_BITS))
    return blocks, intervals


def check_overlaps(rows: list[BaseRow], label: str) -> None:
    systems = sorted({sanitize_identifier(r.system_name) for r in rows})
    for sys_key in systems:
        blocks, intervals = system_local_bits(rows, sys_key)
        if blocks == 0:
            continue
        intervals.sort(key=lambda x: x[0])
        for i in range(1, len(intervals)):
            p_start, p_end, p_row = intervals[i - 1]
            c_start, c_end, c_row = intervals[i]
            if c_start < p_end:
                fail(
                    f"{label}/{sys_key}: пересечение битовых полей: "
                    f"{p_row.name}[{p_start}:{p_end}] и {c_row.name}[{c_start}:{c_end}]"
                )


def system_stats(rows: list[BaseRow], systems: list[str], label: str) -> dict[str, SystemStats]:
    result: dict[str, SystemStats] = {}

    for sys in systems:
        sys_key = sanitize_identifier(sys)
        blocks, intervals = system_local_bits(rows, sys_key)
        bits = blocks * BLOCK_BITS
        result[sys_key] = SystemStats(blocks=blocks, bits=bits)

        if blocks == 0:
            info(f"{label}/{sys_key}: данных нет")
            continue

        occupied = [False] * bits
        for start, end, _ in intervals:
            for b in range(max(0, start), min(bits, end)):
                occupied[b] = True

        free_ranges: list[tuple[int, int]] = []
        idx = 0
        while idx < bits:
            if occupied[idx]:
                idx += 1
                continue
            begin = idx
            while idx < bits and not occupied[idx]:
                idx += 1
            free_ranges.append((begin, idx))

        if free_ranges:
            pretty = ", ".join([f"[{a}:{b})" for a, b in free_ranges])
            info(f"{label}/{sys_key}: незанятые биты: {pretty}")
        else:
            info(f"{label}/{sys_key}: незанятых бит нет")

    return result


def compute_cs_maps(systems: list[str], stats: dict[str, SystemStats]) -> tuple[dict[str, int], int]:
    base_bits: dict[str, int] = {}
    current = 0
    for sys in systems:
        key = sanitize_identifier(sys)
        if key in ("NONE", "CS"):
            continue
        base_bits[key] = current
        current += stats.get(key, SystemStats(0, 0)).bits
    return base_bits, current


def assign_cs_offsets(rows: list[BaseRow], cs_base_bits: dict[str, int]) -> None:
    # local base block index for each system
    min_block_by_system: dict[str, int] = {}
    for row in rows:
        if row.is_placeholder:
            continue
        key = sanitize_identifier(row.system_name)
        if key not in min_block_by_system:
            min_block_by_system[key] = row.msg_block_n
        else:
            min_block_by_system[key] = min(min_block_by_system[key], row.msg_block_n)

    for row in rows:
        key = sanitize_identifier(row.system_name)
        local_min = min_block_by_system.get(key, 0)
        local_bit = (row.msg_block_n - local_min) * BLOCK_BITS + row.msg_block_offset
        base = cs_base_bits.get(key, 0)
        row.msg_cs_offset = base + local_bit
        row.msg_cs_block_n = row.msg_cs_offset // BLOCK_BITS


def make_define(name: str, value: Any, comment: str) -> str:
    prefix = f"#define {name:<34} {str(value):<6} "
    comment_width = LINE_WIDTH - len(prefix) - len("/** ") - len("*/")
    comment_text = c_comment(comment)
    if len(comment_text) < comment_width:
        comment_text = comment_text + (" " * (comment_width - len(comment_text)))
    else:
        comment_text = comment_text[:comment_width]
    return f"{prefix}/** {comment_text}*/"


def make_define_file(items: list[tuple[str, Any, str]], max_name: str, max_value: int) -> str:
    lines: list[str] = [banner("BD.xlsx"), pad140("#pragma once"), ""]
    for name, value, comment in items:
        lines.append(make_define(name, value, comment))
    lines.append("")
    lines.append(make_define(max_name, max_value, ""))
    lines.append("")
    return "\n".join(lines)


def make_sys_list(
    systems: list[str],
    pstats: dict[str, SystemStats],
    cstats: dict[str, SystemStats],
    cs_param_bits: int,
    cs_comm_bits: int,
) -> str:
    items: list[tuple[str, Any, str]] = [
        ("SYS_NONE", 0, "id системы NONE"),
        ("SYS_CS", 1, "id системы CS"),
    ]

    filtered: list[str] = []
    seen: set[str] = set()
    for name in systems:
        key = sanitize_identifier(name)
        if key in ("NONE", "CS") or key in seen:
            continue
        seen.add(key)
        filtered.append(name)

    for idx, name in enumerate(filtered, start=2):
        items.append((f"SYS_{sanitize_identifier(name)}", idx, f"id системы {name}"))

    cs_param_storage = SystemStats(cs_param_bits // BLOCK_BITS, cs_param_bits)
    cs_comm_storage = SystemStats(cs_comm_bits // BLOCK_BITS, cs_comm_bits)
    cs_param_buf = SystemStats(
        max((pstats.get(sanitize_identifier(name), SystemStats(0, 0)).blocks for name in filtered), default=0),
        max((pstats.get(sanitize_identifier(name), SystemStats(0, 0)).bits for name in filtered), default=0),
    )
    cs_comm_buf = SystemStats(
        max((cstats.get(sanitize_identifier(name), SystemStats(0, 0)).blocks for name in filtered), default=0),
        max((cstats.get(sanitize_identifier(name), SystemStats(0, 0)).bits for name in filtered), default=0),
    )

    items.extend(
        [
            ("CS_PARAM_STORAGE_BLOCKS", cs_param_storage.blocks, "CS: storage blocks for parameters"),
            ("CS_PARAM_STORAGE_BITS", cs_param_storage.bits, "CS: storage bit length for parameters"),
            ("CS_COMM_STORAGE_BLOCKS", cs_comm_storage.blocks, "CS: storage blocks for commands"),
            ("CS_COMM_STORAGE_BITS", cs_comm_storage.bits, "CS: storage bit length for commands"),
            ("CS_PARAM_BUF_BLOCKS", cs_param_buf.blocks, "CS: transport buffer blocks for parameters"),
            ("CS_PARAM_BUF_BITS", cs_param_buf.bits, "CS: transport buffer bit length for parameters"),
            ("CS_COMM_BUF_BLOCKS", cs_comm_buf.blocks, "CS: transport buffer blocks for commands"),
            ("CS_COMM_BUF_BITS", cs_comm_buf.bits, "CS: transport buffer bit length for commands"),
        ]
    )

    metric_targets = ["CS", *filtered]
    for name in metric_targets:
        key = sanitize_identifier(name)
        p = pstats.get(key, SystemStats(0, 0))
        c = cstats.get(key, SystemStats(0, 0))

        if key == "CS":
            p = cs_param_storage
            c = cs_comm_storage

        items.extend(
            [
                (f"{key}_PARAM_MSG_BLOCKS", p.blocks, f"{name}: max блоков параметров"),
                (f"{key}_PARAM_MSG_BITS", p.bits, f"{name}: max длина параметров, бит"),
                (f"{key}_COMM_MSG_BLOCKS", c.blocks, f"{name}: max блоков команд"),
                (f"{key}_COMM_MSG_BITS", c.bits, f"{name}: max длина команд, бит"),
            ]
        )

    return make_define_file(items, "SYS_max", len(filtered) + 2)

def make_alg_list(algs: list[str]) -> str:
    items = []
    for idx, name in enumerate(algs):
        comment = "алгоритм обработки" if idx == 0 and sanitize_identifier(name) == "NONE" else name
        items.append((f"ALG_{sanitize_identifier(name)}", idx, comment))
    return make_define_file(items, "ALG_max", len(algs))


def make_alg_tab(algorithms: list[AlgorithmRow]) -> str:
    lines = [
        banner("BD.xlsx/Algorithms"),
        pad140("#pragma once"),
        "",
        pad140('#include "alg_list.hpp"'),
        "",
        pad140("struct S_algtab{"),
        field_line("const char*", "key;", "идентификатор алгоритма"),
        field_line("const char*", "encode_expr;", "формула encode"),
        field_line("const char*", "decode_expr;", "формула decode"),
        field_line("const char*", "condition_expr;", "условие применимости"),
        pad140("};"),
        "",
        pad140("inline S_algtab S_basealgtab[ALG_max]={"),
    ]

    init_rows = ['\t{"NONE", "", "", ""}']
    comments: list[tuple[str, str]] = [("NONE", "reserved")]
    commas: list[bool] = [True]

    for row in algorithms:
        init_rows.append(
            (
                f'\t{{"{sanitize_identifier(row.name)}", '
                f'"{c_comment(row.encode_expr)}", '
                f'"{c_comment(row.decode_expr)}", '
                f'"{c_comment(row.condition_expr)}"}}'
            )
        )
        comments.append(table_comment(row.name, row.decode_expr))
        commas.append(True)

    commas[-1] = False
    lines.extend(align_rows(init_rows, comments, commas))
    lines.append(pad140("};"))
    lines.append("")
    return "\n".join(lines)


def make_type_list(types: list[str]) -> str:
    items = []
    for idx, name in enumerate(types):
        comment = "тип данных" if idx == 0 and sanitize_identifier(name) == "NONE" else name
        items.append((f"TYPE_{sanitize_identifier(name)}", idx, comment))
    return make_define_file(items, "TYPE_max", len(types))


def make_real_type_list() -> str:
    items = [
        ("RTYPE_NONE", 0, "physical type none"),
        ("RTYPE_BOOL", 1, "physical type bool"),
        ("RTYPE_INT8", 2, "physical type int8_t"),
        ("RTYPE_UINT8", 3, "physical type uint8_t"),
        ("RTYPE_INT16", 4, "physical type int16_t"),
        ("RTYPE_UINT16", 5, "physical type uint16_t"),
        ("RTYPE_FLOAT32", 6, "physical type float32_t"),
        ("RTYPE_UFLOAT32", 7, "physical type ufloat32_t"),
    ]
    return make_define_file(items, "RTYPE_max", len(items))


def make_param_comm_list(params: list[ParamRow], commands: list[CommandRow]) -> str:
    param_count = len(params)
    command_count = len(commands)
    total_count = param_count + command_count

    lines: list[str] = [banner("BD.xlsx"), pad140("#pragma once"), ""]
    param_names: set[str] = set()

    for row in params:
        name = sanitize_identifier(row.name)
        lines.append(make_define(name, row.id, row.description))
        param_names.add(name)

    lines.append("")
    for row in commands:
        name = sanitize_identifier(row.name)
        if name in param_names:
            lines.append(pad140(f"#undef {name}"))
        lines.append(make_define(name, param_count + row.id, row.description))
    lines.append("")
    lines.append(make_define("Param_max", param_count, ""))
    lines.append(make_define("Comm_max", command_count, ""))
    lines.append(make_define("Param_Comm_max", total_count, ""))
    lines.append("")

    return "\n".join(lines)


def table_comment(name: str, desc: str) -> tuple[str, str]:
    return (f'"{name}"', f'"{desc}"')


def align_rows(init_rows: list[str], comments: list[tuple[str, str]], commas: list[bool]) -> list[str]:
    rows_with_commas = [f"{r}{',' if c else ''}" for r, c in zip(init_rows, commas)]
    max_len = max((len(r) for r in rows_with_commas), default=0)
    out = []
    for row, (left_comment, right_comment) in zip(rows_with_commas, comments):
        prefix = f"{row:<{max_len}}	"
        comment = make_aligned_comment(left_comment, right_comment, LINE_WIDTH - len(prefix))
        out.append((prefix + comment)[:LINE_WIDTH])
    return out


def make_param_tab(params: list[ParamRow]) -> str:
    lines = [
        banner("BD.xlsx"),
        pad140("#pragma once"),
        pad140("#include <stdint.h>"),
        "",
        pad140('#include "sys_list.hpp"'),
        pad140('#include "type_list.hpp"'),
        pad140('#include "real_type_list.hpp"'),
        pad140('#include "alg_list.hpp"'),
        pad140('#include "param_comm_list.hpp"'),
        "",
        pad140("#define\tTAR_NONE\t0"),
        "",
        pad140("struct S_paramtab{"),
        field_line("char", "sys_id;", "id системы"),
        field_line("char", "type;", "тип данных"),
        field_line("char", "type_real;", "физический тип"),
        field_line("char", "sign;", "TYPE_SIGN / TYPE_UNSIGN"),
        field_line("double", "tar_a;", "тарировка A"),
        field_line("double", "tar_b;", "тарировка B"),
        field_line("double", "tar_c;", "тарировка C"),
        field_line("uint16_t", "msg_offset;", "глобальное смещение в сообщении"),
        field_line("char", "msg_block_offset;", "смещение в блоке данных"),
        field_line("char", "msg_block_n;", "номер блока данных"),
        field_line("uint16_t", "msg_cs_offset;", "глобальное смещение внутри CS"),
        field_line("char", "msg_cs_block_n;", "номер блока внутри CS"),
        field_line("const char*", "key;", "идентификатор параметра"),
        field_line("char", "alg;", "алгоритм обработки"),
        pad140("};"),
        "",
        pad140("inline S_paramtab S_baseparamtab[Param_max]={"),
    ]

    init_rows: list[str] = []
    comments: list[tuple[str, str]] = []
    commas: list[bool] = []
    for idx, r in enumerate(params):
        init_rows.append(
            (
                f"\t{{SYS_{sanitize_identifier(r.system_name)}, TYPE_{sanitize_identifier(r.type_name)}, RTYPE_{sanitize_identifier(r.type_real)}, {sign_macro(r.sign)}, "
                f"{to_cpp_number(r.tar_a)}, {to_cpp_number(r.tar_b)}, {to_cpp_number(r.tar_c)}, {r.msg_offset}, "
                f'{r.msg_block_offset}, {r.msg_block_n}, {r.msg_cs_offset}, {r.msg_cs_block_n}, "{sanitize_identifier(r.name)}", '
                f"ALG_{sanitize_identifier(r.alg_name)}}}"
            )
        )
        comments.append(table_comment(r.name, r.description))
        commas.append(idx != len(params) - 1)

    lines.extend(align_rows(init_rows, comments, commas))
    lines.append(pad140("};"))
    lines.append("")
    return "\n".join(lines)

def make_command_tab(commands: list[CommandRow]) -> str:
    lines = [
        banner("BD.xlsx"),
        pad140("#pragma once"),
        pad140("#include <stdint.h>"),
        "",
        pad140('#include "sys_list.hpp"'),
        pad140('#include "type_list.hpp"'),
        pad140('#include "real_type_list.hpp"'),
        pad140('#include "alg_list.hpp"'),
        pad140('#include "param_comm_list.hpp"'),
        "",
        pad140("#define\tTAR_NONE\t0"),
        "",
        pad140("struct S_commtab{"),
        field_line("char", "sys_id;", "id системы"),
        field_line("char", "type;", "тип данных"),
        field_line("char", "type_real;", "физический тип"),
        field_line("char", "sign;", "TYPE_SIGN / TYPE_UNSIGN"),
        field_line("double", "tar_a;", "тарировка A"),
        field_line("double", "tar_b;", "тарировка B"),
        field_line("double", "tar_c;", "тарировка C"),
        field_line("uint16_t", "msg_offset;", "глобальное смещение в сообщении"),
        field_line("char", "msg_block_offset;", "смещение в блоке данных"),
        field_line("char", "msg_block_n;", "номер блока данных"),
        field_line("uint16_t", "msg_cs_offset;", "глобальное смещение внутри CS"),
        field_line("char", "msg_cs_block_n;", "номер блока внутри CS"),
        field_line("const char*", "key;", "идентификатор команды"),
        field_line("char", "alg;", "алгоритм обработки"),
        pad140("};"),
        "",
        pad140("inline S_commtab S_basecommtab[Comm_max]={"),
    ]

    init_rows: list[str] = []
    comments: list[tuple[str, str]] = []
    commas: list[bool] = []
    for idx, r in enumerate(commands):
        init_rows.append(
            (
                f"\t{{SYS_{sanitize_identifier(r.system_name)}, TYPE_{sanitize_identifier(r.type_name)}, RTYPE_{sanitize_identifier(r.type_real)}, {sign_macro(r.sign)}, "
                f"{to_cpp_number(r.tar_a)}, {to_cpp_number(r.tar_b)}, {to_cpp_number(r.tar_c)}, {r.msg_offset}, "
                f'{r.msg_block_offset}, {r.msg_block_n}, {r.msg_cs_offset}, {r.msg_cs_block_n}, "{sanitize_identifier(r.name)}", '
                f"ALG_{sanitize_identifier(r.alg_name)}}}"
            )
        )
        comments.append(table_comment(r.name, r.description))
        commas.append(idx != len(commands) - 1)

    lines.extend(align_rows(init_rows, comments, commas))
    lines.append(pad140("};"))
    lines.append("")
    return "\n".join(lines)


def system_rows(rows: list[BaseRow], system_key: str) -> list[BaseRow]:
    selected = [r for r in rows if sanitize_identifier(r.system_name) == system_key and not r.is_placeholder]
    return sorted(selected, key=lambda r: r.id)


def local_model_rows(rows: list[ModelRow], system_key: str) -> list[ModelRow]:
    selected = [r for r in rows if sanitize_identifier(r.system_name) == system_key]
    return sorted(selected, key=lambda r: (r.id if r.id is not None else 1 << 30, r.row_num, sanitize_identifier(r.name)))

def resolve_local_model_rows(system_key: str, local_rows: list[ModelRow]) -> list[ModelRow]:
    locals_for_system = local_model_rows(local_rows, system_key)
    if not locals_for_system:
        return []

    used_ids: set[int] = set()
    next_local_id = max([row.id for row in locals_for_system if row.id is not None], default=-1) + 1
    resolved_locals: list[ModelRow] = []

    for row in locals_for_system:
        row_id = row.id
        if row_id is None:
            while next_local_id in used_ids:
                next_local_id += 1
            row_id = next_local_id
            next_local_id += 1
        elif row_id in used_ids:
            fail(f"Local_Params/{system_key}: duplicate local ID={row_id}")

        if row_id < 0:
            fail(f"Local_Params/{system_key}: local ID must be non-negative, got {row_id}")

        used_ids.add(row_id)
        resolved_locals.append(
            ModelRow(
                row_num=row.row_num,
                id=row_id,
                name=row.name,
                system_name=row.system_name,
                type_name=row.type_name,
                type_real=row.type_real,
                sign=row.sign,
                init_value=row.init_value,
            )
        )

    return sorted(resolved_locals, key=lambda r: r.id if r.id is not None else 0)


def as_contiguous_range(ids: list[int], label: str) -> tuple[int, int]:
    if not ids:
        return 0, 0

    ordered = sorted(ids)
    base = ordered[0]
    expected = list(range(base, base + len(ordered)))
    if ordered != expected:
        warn(f"{label}: IDs are not contiguous, range will use base={base} count={len(ordered)}")
    return base, len(ordered)


def model_param_var_name(row: ModelLike) -> str:
    return f"P_{sanitize_identifier(row.name)}"


def model_param_wire_bits(row: ModelLike) -> int:
    return type_size_bits(row.type_name)


def model_param_runtime_id(row: ModelLike) -> int:
    if isinstance(row, ModelRow):
        if row.id is None:
            fail(f"Local model row {row.name} has unresolved ID")
        return LOCAL_PARAM_ID_BASE + row.id
    return row.id


def model_param_cpp_type(row: ModelLike) -> str:
    kind = sanitize_identifier(row.type_real)

    if kind == "BOOL":
        return "std::uint8_t"
    if kind == "INT8":
        return "std::int8_t"
    if kind == "UINT8":
        return "std::uint8_t"
    if kind == "INT16":
        return "std::int16_t"
    if kind == "UINT16":
        return "std::uint16_t"
    if kind in ("FLOAT32", "UFLOAT32"):
        return "float"

    fail(f"Unsupported physical type for ID={row.id} name={row.name}: {row.type_real}")
    return "std::uint8_t"


def model_param_real_enum(row: ModelLike) -> str:
    kind = sanitize_identifier(row.type_real)

    if kind == "BOOL":
        return "sim_base::physical_type_t::boolean"
    if kind == "INT8":
        return "sim_base::physical_type_t::int8"
    if kind == "UINT8":
        return "sim_base::physical_type_t::uint8"
    if kind == "INT16":
        return "sim_base::physical_type_t::int16"
    if kind == "UINT16":
        return "sim_base::physical_type_t::uint16"
    if kind == "FLOAT32":
        return "sim_base::physical_type_t::float32"
    if kind == "UFLOAT32":
        return "sim_base::physical_type_t::ufloat32"

    return "sim_base::physical_type_t::uint8"


def model_param_init_literal(row: ModelLike) -> str:
    return to_cpp_number(row.init_value)


def make_sys_id_range(systems: list[str], params: list[ParamRow], commands: list[CommandRow]) -> str:
    lines: list[str] = [banner("BD.xlsx"), pad140("#pragma once"), ""]
    param_count = len(params)

    system_keys: list[str] = []
    for name in systems:
        key = sanitize_identifier(name)
        if key in ("NONE", "CS") or key in system_keys:
            continue
        system_keys.append(key)

    for key in system_keys:
        p_rows = system_rows(params, key)
        p_ids = [r.id for r in p_rows]
        base, count = as_contiguous_range(p_ids, f"PARAM/{key}")
        lines.append(make_define(f"{key}_BASE", f"{base}u", f"{key}: parameter ID base"))
        lines.append(make_define(f"{key}_COUNT", f"{count}u", f"{key}: parameter ID count"))
        lines.append("")

    for key in system_keys:
        c_rows = system_rows(commands, key)
        c_ids = [param_count + r.id for r in c_rows]
        base, count = as_contiguous_range(c_ids, f"COMM/{key}")
        lines.append(make_define(f"C{key}_BASE", f"{base}u", f"{key}: command ID base"))
        lines.append(make_define(f"C{key}_COUNT", f"{count}u", f"{key}: command ID count"))
        lines.append("")

    return "\n".join(lines)


def make_local_param_list(local_params: list[ModelRow]) -> str:
    items: list[tuple[str, Any, str]] = [
        ("LOCAL_PARAM_ID_BASE", f"0x{LOCAL_PARAM_ID_BASE:04X}u", "base offset for model-local ISIM identifiers"),
    ]

    for row in sorted(local_params, key=lambda r: (r.id if r.id is not None else 1 << 30, sanitize_identifier(r.name))):
        if row.id is None:
            fail(f"Local_Params: unresolved ID for {row.name}")
        items.append((sanitize_identifier(row.name), model_param_runtime_id(row), f"{row.system_name}: local model parameter"))

    return make_define_file(items, "LOCAL_PARAM_max", len(local_params))


def make_model_param_array(array_name: str, count_name: str, rows: list[ModelLike]) -> list[str]:
    lines = [pad140(f"inline std::array<sim_base::param_entry_t, {count_name}> {array_name} = {{")]

    for idx, row in enumerate(rows):
        tail = "," if idx != (len(rows) - 1) else ""
        bits = model_param_wire_bits(row)
        real_kind = model_param_real_enum(row)
        key = sanitize_identifier(row.name)
        runtime_id = model_param_runtime_id(row)
        name = model_param_var_name(row)
        lines.append(
            pad140(
                f'  sim_base::param_entry_t{{{runtime_id}, "{key}", {bits}U, {real_kind}, static_cast<void*>(&{name}), '
                f'{model_param_init_literal(row)}}}{tail}'
            )
        )

    lines.append(pad140("};"))
    return lines


def make_sim_model_base_hpp(namespace_name: str, params_rows: list[BaseRow], local_rows: list[ModelRow]) -> str:
    all_rows: list[ModelLike] = [*params_rows, *local_rows]
    param_vars = [model_param_var_name(r) for r in all_rows]

    lines = [
        banner("BD.xlsx"),
        pad140("#pragma once"),
        "",
        pad140('#include "sim_base.hpp"'),
        "",
        pad140(f"namespace {namespace_name} {{"),
        "",
        pad140(f"constexpr std::size_t k_param_count = {len(params_rows)}U;"),
        pad140(f"constexpr std::size_t k_local_param_count = {len(local_rows)}U;"),
        pad140(f"constexpr std::size_t k_isim_param_count = {len(all_rows)}U;"),
        "",
    ]

    for row, name in zip(all_rows, param_vars):
        lines.append(pad140(f"inline {model_param_cpp_type(row)} {name} = {model_param_init_literal(row)};"))

    if param_vars:
        lines.append("")

    lines.extend(make_model_param_array("PARAMS", "k_param_count", params_rows))
    lines.append("")
    lines.extend(make_model_param_array("LOCAL_PARAMS", "k_local_param_count", local_rows))
    lines.append("")
    lines.extend(make_model_param_array("ISIM_PARAMS", "k_isim_param_count", all_rows))
    lines.append("")
    lines.append(pad140("inline void init_defaults() {"))
    lines.append(pad140("  sim_base::init_params(PARAMS);"))
    lines.append(pad140("  sim_base::init_params(LOCAL_PARAMS);"))
    lines.append(pad140("}"))
    lines.append("")
    lines.append(pad140(f"}}  // namespace {namespace_name}"))
    lines.append("")
    return "\n".join(lines)


def make_sim_base_files(systems: list[str],
                        params: list[ParamRow],
                        local_params: list[ModelRow]) -> dict[Path, str]:
    files: dict[Path, str] = {}

    system_keys: list[str] = []
    for name in systems:
        key = sanitize_identifier(name)
        if key == "NONE" or key in system_keys:
            continue
        system_keys.append(key)

    for row in local_params:
        key = sanitize_identifier(row.system_name)
        if key == "NONE" or key in system_keys:
            continue
        system_keys.append(key)

    for key in system_keys:
        regular_rows = system_rows(params, key)
        local_rows = resolve_local_model_rows(key, local_params)
        if not regular_rows and not local_rows:
            continue

        sys_lc = key.lower()
        namespace_name = f"sim_{sys_lc}_base"
        header_name = f"sim_{sys_lc}_base.hpp"
        files[SIM_BASE_INCLUDE_DIR / header_name] = make_sim_model_base_hpp(namespace_name, regular_rows, local_rows)

    return files


def write_file(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8", newline="\n")


def main() -> None:
    wb = load_data_book(EXCEL_PATH)

    params_raw = parse_sheet(wb, "Params", "Params")
    algorithms = parse_algorithm_sheet(wb, "Algorithms")
    commands_raw = parse_sheet(wb, "Commands", "Commands")
    local_model_rows = parse_model_sheet(wb, "Local_Params", "Local_Params")

    params = cast_params(fill_id_gaps(params_raw, "Params"))
    commands = cast_commands(fill_id_gaps(commands_raw, "Commands"))

    validate_block_layout(params, "Params")
    validate_block_layout(commands, "Commands")
    check_overlaps(params, "Params")
    check_overlaps(commands, "Commands")
    validate_algorithm_references(params, algorithms, "Params")
    validate_algorithm_references(commands, algorithms, "Commands")
    validate_algorithm_conditions(params, "Params")
    validate_algorithm_conditions(commands, "Commands")
    report_quantization_errors(params, "Params")
    report_quantization_errors(commands, "Commands")

    systems = collect_systems(params, commands)
    algs = collect_algs(algorithms)
    types = collect_types(params, commands)

    pstats = system_stats(params, systems, "Params")
    cstats = system_stats(commands, systems, "Commands")

    cs_param_base, cs_param_bits = compute_cs_maps(systems, pstats)
    cs_comm_base, cs_comm_bits = compute_cs_maps(systems, cstats)

    info(f"CS/PARAM: суммарно блоков={cs_param_bits // BLOCK_BITS}, бит={cs_param_bits}")
    info(f"CS/COMM: суммарно блоков={cs_comm_bits // BLOCK_BITS}, бит={cs_comm_bits}")

    assign_cs_offsets(params, cs_param_base)
    assign_cs_offsets(commands, cs_comm_base)

    ensure_dir(OUTPUT_DIR)
    ensure_dir(SIM_BASE_INCLUDE_DIR)

    files = {
        "sys_list.hpp": make_sys_list(systems, pstats, cstats, cs_param_bits, cs_comm_bits),
        "sys_id_range.hpp": make_sys_id_range(systems, params, commands),
        "alg_list.hpp": make_alg_list(algs),
        "alg_tab.hpp": make_alg_tab(algorithms),
        "local_param_list.hpp": make_local_param_list(local_model_rows),
        "type_list.hpp": make_type_list(types),
        "real_type_list.hpp": make_real_type_list(),
        "param_comm_list.hpp": make_param_comm_list(params, commands),
        "param_tab.hpp": make_param_tab(params),
        "command_tab.hpp": make_command_tab(commands),
    }

    for legacy_name in ("param_list.hpp", "command_list.hpp"):
        legacy_path = OUTPUT_DIR / legacy_name
        if legacy_path.exists():
            legacy_path.unlink()

    for fname, content in files.items():
        write_file(OUTPUT_DIR / fname, content)

    for stale in SIM_BASE_INCLUDE_DIR.glob("sim_*_base.hpp"):
        stale.unlink()

    sim_files = make_sim_base_files(systems, params, local_model_rows)
    for path, content in sim_files.items():
        write_file(path, content)

    print(f"Generated {len(files)} files in: {OUTPUT_DIR}")
    print(f"Generated {len(sim_files)} files in: {SIM_BASE_INCLUDE_DIR.parent}")
    print("Done!")


if __name__ == "__main__":
    main()
